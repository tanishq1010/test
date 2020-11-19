import requests
import pandas as pd
from openpyxl import Workbook, load_workbook
import os


class Source(object):
    def __init__(self):
        super(Source, self).__init__()
        self.headers = {
            'Connection': 'keep-alive',
            'Accept': '*/*',
            'Content-Type': 'application/json',
        }
        self.host = 'https://preprodms.embibe.com'

    def callAPI(self, url, payload, method):
        self.headers[
            'embibe-token'] = 'eyJhbGciOiJIUzUxMiJ9.eyJpZCI6MTExNzE5NTMxNiwiZW1haWwiOiJkaGFucmFqMTk5NEBnbWFpbC5jb20iLCJvcmdhbml6YXRpb25faWQiOm51bGwsImlzX2d1ZXN0IjpmYWxzZSwicm9sZSI6InN0dWRlbnQiLCJ0aW1lX3N0YW1wIjoiMjAyMC0wNy0xMFQwNjo1MDowMi43NDhaIn0.y_OLKsIY2F8CdSDDzcKLvO_EpBFyQyvdQWlpS9wqzh6_Ap1xTyKPnKpXl_QBvzXZMoskcXn0YLNN_0l9Z2EFpw'
        response = requests.request(method, self.host + url, headers=self.headers, data=payload)
        if response.status_code != 200:
            print(url + ' - ' + str(response.content))
            return None
        return response

    def main(self,goal,exam):
        # workbook = Workbook()
        # sheet = workbook.active
        # sheet["A1"] = "Goal"
        # sheet["B1"] = "Exam"
        # sheet["C1"] = "Bundle_id"
        # sheet["D1"] = "Title"
        # sheet["E1"] = "Learnpath_Name"

        # workbook.save(filename="data_exam.xlsx")
        # goal = "CBSE"
        # exam = "10th CBSE"
        # query = "MATHEMATICS"
        var3 = exam
        for i in range(0, len(var3)):
            if var3[i] == " ":
                var3 = var3[:i] + "-" + var3[i + 1:]
        response1 = self.callAPI(
            "/content_ms_fiber/v2/learning-maps/" + str(goal.lower()) + "/" + str(
                var3.lower()) + "/fetch-test-exam-test-data",
            "{}",
            'GET')

        print(response1.status_code)
        # print(response1.json())
        # wb = load_workbook("data_exam.xlsx")
        # sheet = wb["Sheet"]
        home_data = []
        for test in response1.json()["results"]["test_list"]:

            # print(value["test_list"])
            # for testlist in value["test_list"]:
            try:
                home = []
                home.append(str(goal.upper()))
                home.append(str(exam.upper()))
                home.append(str(test["bundle_id"]))
                home.append(str(test["title"]).upper())
                home.append(str(test["learnpath_name"]).upper())
                home_data.append(home)

            except Exception as e:
                print(e)

        # wb.save(filename="data_exam.csv")

        df = pd.DataFrame(home_data, columns=['Goal',"Exam","Bundle_id","Title","Learnpath_Name"])

        return df


def remove_space(string): 
    return string.replace(" ", "")    


def test_data_extractor(goal,exam,ID):
    src = Source()
    exam=exam.upper()
    goal=goal.upper()

    if os.path.exists(goal +"_"+ remove_space(str(exam)) + "_cg_test_data.csv" ):
        print("\tFile: "+  goal +"_" +remove_space(str(exam)) + "_cg_test_data.csv found. Reading....")
        df = pd.read_csv(goal +"_"+ remove_space(str(exam)) + "_cg_test_data.csv")
        df=df.loc[df['Bundle_id'].str.contains(ID)]
        if len(df) == 0:
            return "NA"

        if len(df[df['Exam'].str.contains(exam)]) > 0:
            return "Yes"
        else:
            return "No"

    else:
        print("\tFile: "+ goal +"_" +remove_space(str(exam)) + "_cg_test_data.csv not found.")
        df = src.main(goal,exam)
        df.to_csv(goal +"_"+ remove_space(str(exam)) + "_cg_test_data.csv",index=False)
        df = pd.read_csv(goal +"_"+ remove_space(str(exam)) + "_cg_test_data.csv")
        
        df=df.loc[df['Bundle_id'].str.contains(ID)]
        if len(df) == 0:
            return "NA"

        if len(df[df['Exam'].str.contains(exam)]) > 0:
            return "Yes"
        else:
            return "No"

